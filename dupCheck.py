#####################################################################################################################################################################################################
    #Duplicate Records Check
    #Input : .xlsx data file to be loaded into CAMP
    #Output: Indicating if there are any duplicate of EntityName/Country Combination and the number of duplicates.
#####################################################################################################################################################################################################


from openpyxl import Workbook, load_workbook
import argparse
import sys


def validateHeader(ws):
    """Validate that entity name and country in the correct columns"""
    returnValue = False
    e = ws.cell(row=1, column=1).value
    c = ws.cell(row=1, column=7).value
    #print(e)
    #print(c)
    if (e == "Entity Name" and c == "Country"):
        returnValue = True
    return returnValue

def getNameCountry(ws, inputRow):
    """given a row number returns the name and county in that row"""
    #expects column 1 to be Country Name
    #expects column 7 to be Country
    entityName = ws.cell(row=inputRow, column=1).value
    countryName = ws.cell(row=inputRow, column=7).value
    return ((entityName, countryName))

def getRows(ws, entityName, country):
    """given an entity name and country returns a list of rows where that combination appears"""
    #print(entityName)
    returnRows = []
    for r in range(2,ws.max_row+1):
        if (ws.cell(row=r,column=1).value == entityName) and (ws.cell(row=r,column=7).value == country):
            returnRows.append(r)
    return returnRows

def runDupCheck(ws):
    entityList = []

    for n in range(1, ws.max_row):
        entityList.append(getNameCountry(ws, n))

    res = {}
    for obj in entityList:
        if obj not in res:
            res[obj] = 0
        else: 
            res[obj] += 1

    if(validateHeader(ws)):
        print("Header Validation : OK")
    else:
        print("Header Validation : \033[1;31m Fail \033[0;0m")
        print("\tExpected Column A : Entity Name")
        print("\tExpected Column G : Country")
    

    print("Rows Checked: " + str(len(entityList)+1))

    noDuplicatesFound = True
    
    for ent in res:
        if res[ent] > 0:
            noDuplicatesFound = False
            entityName = ent[0]
            country = ent[1]
            dupRows = getRows(ws, entityName,country)
            print("Duplicate Validation : \033[1;31m Fail \033[0;0m")
            print("\tThe entity " + entityName + "/" + country + " has duplicate entries in the file.  See rows: " + str(dupRows))

    if (noDuplicatesFound):
        print("No duplicates entity name/country combinations found.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", help="a .xlsx file to be loaded into CAMP")
    args = parser.parse_args()

    filename = args.filename

    #load workbook
    print("\n")
    print("Checking: " + filename)
    wb = load_workbook(filename)

    #activate sheet
    ws = wb.active

    #get max rows
    max_row = ws.max_row

    runDupCheck(ws)

if (__name__ == "__main__"):
    main()

